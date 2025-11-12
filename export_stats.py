import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import io

def generate_stats_excel(db_path):
    """
    Gera um arquivo Excel com estatísticas completas dos pedidos
    """
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Criar workbook
    wb = Workbook()
    
    # ===== SHEET 1: Estatísticas Gerais =====
    ws1 = wb.active
    ws1.title = "Estatísticas Gerais"
    
    # Cabeçalho
    ws1['A1'] = 'Estatística'
    ws1['B1'] = 'Valor'
    ws1['A1'].font = Font(bold=True, size=12)
    ws1['B1'].font = Font(bold=True, size=12)
    ws1['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws1['B1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    # Total de pedidos
    cursor.execute("SELECT COUNT(*) as total FROM requests")
    total = cursor.fetchone()['total']
    
    # Por status
    cursor.execute("SELECT status, COUNT(*) as count FROM requests GROUP BY status")
    status_counts = {row['status']: row['count'] for row in cursor.fetchall()}
    
    # Por tipo
    cursor.execute("SELECT request_type, COUNT(*) as count FROM requests GROUP BY request_type")
    type_counts = {row['request_type']: row['count'] for row in cursor.fetchall()}
    
    # Por período
    cursor.execute("SELECT period, COUNT(*) as count FROM requests GROUP BY period")
    period_counts = {row['period']: row['count'] for row in cursor.fetchall()}
    
    # Preencher dados
    row = 2
    ws1[f'A{row}'] = 'Total de Pedidos'
    ws1[f'B{row}'] = total
    row += 1
    
    ws1[f'A{row}'] = ''
    row += 1
    ws1[f'A{row}'] = 'Por Status:'
    ws1[f'A{row}'].font = Font(bold=True)
    row += 1
    
    for status, count in status_counts.items():
        ws1[f'A{row}'] = f'  {status}'
        ws1[f'B{row}'] = count
        row += 1
    
    row += 1
    ws1[f'A{row}'] = 'Por Tipo:'
    ws1[f'A{row}'].font = Font(bold=True)
    row += 1
    
    for tipo, count in type_counts.items():
        ws1[f'A{row}'] = f'  {tipo}'
        ws1[f'B{row}'] = count
        row += 1
    
    row += 1
    ws1[f'A{row}'] = 'Por Período:'
    ws1[f'A{row}'].font = Font(bold=True)
    row += 1
    
    for period, count in period_counts.items():
        ws1[f'A{row}'] = f'  {period}'
        ws1[f'B{row}'] = count
        row += 1
    
    # Ajustar largura das colunas
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 15
    
    # ===== SHEET 2: Top Lojas =====
    ws2 = wb.create_sheet("Top Lojas")
    
    # Cabeçalho
    headers = ['Posição', 'Loja', 'Total', 'Aprovados', 'Rejeitados', 'Cancelados', 'Pendentes']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Dados
    cursor.execute("""
        SELECT 
            u.shop_name,
            COUNT(*) as total,
            SUM(CASE WHEN r.status = 'Aprovado' THEN 1 ELSE 0 END) as aprovados,
            SUM(CASE WHEN r.status = 'Rejeitado' THEN 1 ELSE 0 END) as rejeitados,
            SUM(CASE WHEN r.status = 'Cancelado' THEN 1 ELSE 0 END) as cancelados,
            SUM(CASE WHEN r.status = 'Pendente' THEN 1 ELSE 0 END) as pendentes
        FROM requests r
        JOIN users u ON r.shop_telegram_id = u.telegram_id
        GROUP BY u.shop_name
        ORDER BY total DESC
        LIMIT 10
    """)
    
    for idx, row_data in enumerate(cursor.fetchall(), 2):
        ws2.cell(row=idx, column=1, value=idx-1)
        ws2.cell(row=idx, column=2, value=row_data['shop_name'])
        ws2.cell(row=idx, column=3, value=row_data['total'])
        ws2.cell(row=idx, column=4, value=row_data['aprovados'])
        ws2.cell(row=idx, column=5, value=row_data['rejeitados'])
        ws2.cell(row=idx, column=6, value=row_data['cancelados'])
        ws2.cell(row=idx, column=7, value=row_data['pendentes'])
    
    # Ajustar largura das colunas
    for col in range(1, 8):
        ws2.column_dimensions[chr(64+col)].width = 15
    ws2.column_dimensions['B'].width = 25
    
    # ===== SHEET 3: Histórico Completo =====
    ws3 = wb.create_sheet("Histórico Completo")
    
    # Cabeçalho
    headers = ['ID', 'Loja', 'Tipo', 'Data', 'Período', 'Status', 'Observações', 'Criado', 'Processado']
    for col, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Dados
    cursor.execute("""
        SELECT 
            r.id,
            u.shop_name,
            r.request_type,
            r.start_date,
            r.period,
            r.status,
            r.observations,
            r.created_at,
            r.processed_at
        FROM requests r
        JOIN users u ON r.shop_telegram_id = u.telegram_id
        ORDER BY r.created_at DESC
    """)
    
    for idx, row_data in enumerate(cursor.fetchall(), 2):
        ws3.cell(row=idx, column=1, value=row_data['id'])
        ws3.cell(row=idx, column=2, value=row_data['shop_name'])
        ws3.cell(row=idx, column=3, value=row_data['request_type'])
        ws3.cell(row=idx, column=4, value=row_data['start_date'])
        ws3.cell(row=idx, column=5, value=row_data['period'])
        ws3.cell(row=idx, column=6, value=row_data['status'])
        ws3.cell(row=idx, column=7, value=row_data['observations'] or '')
        ws3.cell(row=idx, column=8, value=row_data['created_at'])
        ws3.cell(row=idx, column=9, value=row_data['processed_at'] or '')
    
    # Ajustar largura das colunas
    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 25
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 12
    ws3.column_dimensions['E'].width = 12
    ws3.column_dimensions['F'].width = 12
    ws3.column_dimensions['G'].width = 30
    ws3.column_dimensions['H'].width = 18
    ws3.column_dimensions['I'].width = 18
    
    conn.close()
    
    # Salvar em memória
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer
